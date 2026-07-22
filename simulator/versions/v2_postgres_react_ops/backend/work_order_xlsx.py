from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from incident_document_api_models import WorkOrderChecklistItem, WorkOrderStructuredContent


_TEMPLATE_PATH = Path(__file__).with_name("assets") / "work_order_template.xlsx"
_SITE_CHECK_SHEET = "01_현장확인"
_PUMP_CHECK_SHEET = "03_펌프점검"
_HEAT_EXCHANGER_CHECK_SHEET = "04_열교환기점검"


def render_work_order_xlsx(
    content: WorkOrderStructuredContent,
    *,
    status_label: str | None = None,
) -> bytes:
    """Render a field-facing work order workbook without AI-only interpretation."""
    workbook = load_workbook(_TEMPLATE_PATH)
    _set_visible_sheets(workbook, content)
    _populate_site_check(workbook[_SITE_CHECK_SHEET], content, status_label=status_label)
    _configure_site_check_print(workbook[_SITE_CHECK_SHEET])
    equipment_sheet_name = _equipment_sheet_name(content)
    if equipment_sheet_name is not None:
        _populate_equipment_checklist(workbook[equipment_sheet_name], content)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _set_visible_sheets(workbook, content: WorkOrderStructuredContent) -> None:
    equipment_sheet_name = _equipment_sheet_name(content)
    visible = {_SITE_CHECK_SHEET}
    if equipment_sheet_name is not None:
        visible.add(equipment_sheet_name)
    for worksheet in workbook.worksheets:
        worksheet.sheet_state = "visible" if worksheet.title in visible else "hidden"
    workbook.active = workbook.sheetnames.index(_SITE_CHECK_SHEET)


def _equipment_sheet_name(content: WorkOrderStructuredContent) -> str | None:
    equipment_type = content.header.equipment_type
    if "열교환" in equipment_type:
        return _HEAT_EXCHANGER_CHECK_SHEET
    if "펌프" in equipment_type:
        return _PUMP_CHECK_SHEET
    return None


def _populate_site_check(
    sheet,
    content: WorkOrderStructuredContent,
    *,
    status_label: str | None,
) -> None:
    header = content.header
    room_name = _mechanical_room_name(header.mechanical_room)
    sheet["A3"] = None
    sheet.row_dimensions[3].hidden = True
    sheet["B5"] = header.document_number
    sheet["D5"] = header.issued_at.strftime("%Y-%m-%d %H:%M")
    sheet["F5"] = status_label or header.status
    sheet["B6"] = header.target_building
    sheet["D6"] = room_name
    sheet["F6"] = header.equipment_type
    sheet["B7"] = header.issue_reason
    sheet["F7"] = header.work_type
    _remove_deprecated_header_fields(sheet)
    sheet["A10"] = f"{content.purpose}\n\n{content.risk_and_evidence}"
    sheet["A10"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[7].height = 52
    sheet.row_dimensions[10].height = 96
    sheet.row_dimensions[11].height = 96
    _populate_restrictions(sheet, content)
    _populate_site_check_rows(sheet, content.checklist or content.commissioning_checklist)
    sheet["A29"] = content.outcome_and_followup
    sheet["A29"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[29].height = 48


def _remove_deprecated_header_fields(sheet) -> None:
    for row_index in range(5, 8):
        sheet[f"G{row_index}"] = None
        sheet[f"H{row_index}"] = None
        cell_range = f"F{row_index}:H{row_index}"
        if cell_range not in {str(existing) for existing in sheet.merged_cells.ranges}:
            sheet.merge_cells(cell_range)


def _populate_restrictions(sheet, content: WorkOrderStructuredContent) -> None:
    target_cells = (("A14", "B14"), ("C14", "D14"), ("E14", "F14"), ("A15", "B15"), ("C15", "D15"), ("E15", "F15"))
    items = content.restriction_or_prep_checklist[: len(target_cells)]
    for checkbox_cell, label_cell in target_cells:
        sheet[checkbox_cell] = None
        sheet[label_cell] = None
    for (checkbox_cell, label_cell), item in zip(target_cells, items, strict=False):
        sheet[checkbox_cell] = "□"
        sheet[label_cell] = item.label
        sheet[label_cell].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[14].height = 44
    sheet.row_dimensions[15].height = 44


def _populate_site_check_rows(sheet, checklist: tuple[WorkOrderChecklistItem, ...]) -> None:
    for cell_range in ("F18:G18", *(f"F{row}:G{row}" for row in range(19, 25))):
        if cell_range not in {str(existing) for existing in sheet.merged_cells.ranges}:
            sheet.merge_cells(cell_range)
    sheet["F18"] = "측정값/현상"
    for row_index in range(19, 25):
        for column in "ABCDEFGH":
            if column != "G":
                sheet[f"{column}{row_index}"] = None
    for row_index, item in enumerate(checklist[:6], start=19):
        sheet[f"A{row_index}"] = item.seq
        sheet[f"B{row_index}"] = item.instrument_or_target
        sheet[f"C{row_index}"] = item.check_or_task_action
        sheet[f"D{row_index}"] = item.pass_fail_criteria or item.completion_condition or ""


def _configure_site_check_print(sheet) -> None:
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _populate_equipment_checklist(sheet, content: WorkOrderStructuredContent) -> None:
    header = content.header
    sheet["B3"] = ""
    sheet["D3"] = header.issued_at.strftime("%Y-%m-%d")
    sheet["F3"] = _mechanical_room_name(header.mechanical_room)
    sheet["H3"] = ""
    sheet["B4"] = header.equipment_type
    sheet["D4"] = ""
    sheet["F4"] = ""
    sheet["H4"] = header.document_number
    for row_index in range(7, 18):
        for column in "EFGH":
            sheet[f"{column}{row_index}"] = ""


def _mechanical_room_name(value: str | None) -> str:
    if value is None or not value.strip():
        return "기계실"
    normalized = value.strip()
    return normalized if normalized.startswith("기계실") else f"기계실 {normalized}"
