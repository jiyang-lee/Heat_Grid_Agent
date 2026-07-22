from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "simulator" / "versions" / "v2_postgres_react_ops" / "backend"
sys.path.insert(0, str(BACKEND))


def test_site_check_export_contains_only_field_context_and_blank_record_cells() -> None:
    from incident_document_api_models import (
        BooleanChecklistItem,
        SAFETY_PERMIT_QUESTIONS,
        SafetyPermitPrecheck,
        SafetyPermitQuestion,
        WorkOrderChecklistItem,
        WorkOrderHeader,
        WorkOrderStructuredContent,
    )
    from work_order_xlsx import render_work_order_xlsx

    content = WorkOrderStructuredContent(
        work_order_kind="site_check",
        header=WorkOrderHeader(
            document_number="WO-TEST-1",
            issued_at="2026-07-22T09:19:00Z",
            priority="low",
            target_building="해오름 아파트",
            mechanical_room="1",
            equipment_type="순환펌프",
            work_type="현장 확인",
            issue_reason="공급온도 저하와 순환 유량 급변",
            status="검토 중",
        ),
        purpose="순환 유량 변화가 확인되어 순환펌프와 관련 계기의 운전 상태를 점검합니다.",
        risk_and_evidence="순환 유량이 118 m³/h에서 86 m³/h로 낮아졌습니다. 현장에서 유량계 표시값과 펌프 운전음을 함께 확인합니다.",
        restriction_or_prep_checklist=tuple(
            BooleanChecklistItem(label=label)
            for label in (
                "설비 정지·분해 금지",
                "밸브 및 제어 설정값 임의 변경 금지",
                "센서 탈거·전기반 개방 금지",
                "고온 배관·회전체 직접 접촉 금지",
                "누수·이상음·과도한 진동 발견 시 즉시 작업 중지",
                "이상 발견 시 현장 관리자에게 보고",
            )
        ),
        checklist=(
            WorkOrderChecklistItem(
                seq=1,
                instrument_or_target="순환펌프",
                check_or_task_action="이상 소음과 진동을 확인",
                pass_fail_criteria="이상 소음과 진동이 없음",
            ),
        ),
        outcome_and_followup="확인 결과에 이상이 있으면 현장 관리자에게 보고하고 후속 정비 여부를 결정합니다.",
        safety_permit_precheck=SafetyPermitPrecheck(
            questions=tuple(
                SafetyPermitQuestion(question=question, applicable=False)
                for question in SAFETY_PERMIT_QUESTIONS
            ),
            permit_required=False,
        ),
    )

    workbook = load_workbook(BytesIO(render_work_order_xlsx(content, status_label="최종 승인")))
    sheet = workbook["01_현장확인"]

    assert sheet["B5"].value == "WO-TEST-1"
    assert sheet["F5"].value == "최종 승인"
    assert sheet["B6"].value == "해오름 아파트"
    assert sheet["D6"].value == "기계실 1"
    assert sheet["B7"].value == "공급온도 저하와 순환 유량 급변"
    assert sheet["F6"].value == "순환펌프"
    assert sheet["F7"].value == "현장 확인"
    assert [sheet[f"G{row}"].value for row in range(5, 8)] == [None, None, None]
    assert [sheet[f"H{row}"].value for row in range(5, 8)] == [None, None, None]
    assert {"F5:H5", "F6:H6", "F7:H7"}.issubset(
        {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    )
    header_text = tuple(
        cell.value
        for row in sheet.iter_rows(min_row=5, max_row=7)
        for cell in row
        if cell.value
    )
    assert "우선순위" not in header_text
    assert "작업기한" not in header_text
    assert "예상 소요" not in header_text
    assert sheet["A3"].value is None
    assert sheet.row_dimensions[3].hidden is True
    assert sheet["A10"].value == f"{content.purpose}\n\n{content.risk_and_evidence}"
    assert [sheet[cell].value for cell in ("B14", "D14", "F14", "B15", "D15", "F15")] == [
        item.label for item in content.restriction_or_prep_checklist
    ]
    assert sheet["A29"].value == content.outcome_and_followup
    assert sheet["B19"].value == "순환펌프"
    assert sheet["C19"].value == "이상 소음과 진동을 확인"
    assert sheet["D19"].value == "이상 소음과 진동이 없음"
    assert sheet["E19"].value is None
    assert sheet["F19"].value is None
    assert sheet["H19"].value is None
    assert "F18:G18" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    assert sheet["F18"].value == "측정값/현상"
    assert "사진" not in tuple(cell.value for row in sheet.iter_rows(min_row=18, max_row=18) for cell in row if cell.value)
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 1
    assert workbook["03_펌프점검"].sheet_state == "visible"
    assert workbook["04_열교환기점검"].sheet_state == "hidden"


def test_heat_exchanger_export_shows_only_heat_exchanger_sheet() -> None:
    from incident_document_api_models import (
        SAFETY_PERMIT_QUESTIONS,
        SafetyPermitPrecheck,
        SafetyPermitQuestion,
        WorkOrderHeader,
        WorkOrderStructuredContent,
    )
    from work_order_xlsx import render_work_order_xlsx

    content = WorkOrderStructuredContent(
        work_order_kind="site_check",
        header=WorkOrderHeader(
            document_number="WO-TEST-30",
            issued_at="2026-07-22T09:19:00Z",
            priority="high",
            target_building="테스트 단지",
            mechanical_room="30",
            equipment_type="열교환기",
            work_type="현장 확인",
        ),
        purpose="환수온도 저하에 따라 열교환기 상태를 확인합니다.",
        risk_and_evidence="환수온도가 42.2°C에서 34.1°C로 낮아졌습니다.",
        outcome_and_followup="결과를 기록하고 이상이 있으면 보고합니다.",
        safety_permit_precheck=SafetyPermitPrecheck(
            questions=tuple(
                SafetyPermitQuestion(question=question, applicable=False)
                for question in SAFETY_PERMIT_QUESTIONS
            ),
            permit_required=False,
        ),
    )

    workbook = load_workbook(BytesIO(render_work_order_xlsx(content)))

    assert workbook["03_펌프점검"].sheet_state == "hidden"
    assert workbook["04_열교환기점검"].sheet_state == "visible"


def test_unknown_equipment_export_hides_equipment_specific_sheets() -> None:
    from incident_document_api_models import (
        SAFETY_PERMIT_QUESTIONS,
        SafetyPermitPrecheck,
        SafetyPermitQuestion,
        WorkOrderHeader,
        WorkOrderStructuredContent,
    )
    from work_order_xlsx import render_work_order_xlsx

    content = WorkOrderStructuredContent(
        work_order_kind="site_check",
        header=WorkOrderHeader(
            document_number="WO-TEST-UNKNOWN",
            issued_at="2026-07-22T09:19:00Z",
            priority="medium",
            target_building="테스트 단지",
            mechanical_room="99",
            equipment_type="대상 계통",
            work_type="현장 확인",
        ),
        purpose="대상 계통의 현장 상태를 확인합니다.",
        risk_and_evidence="현장 계기값과 외관 상태를 확인합니다.",
        outcome_and_followup="결과를 기록하고 이상이 있으면 보고합니다.",
        safety_permit_precheck=SafetyPermitPrecheck(
            questions=tuple(
                SafetyPermitQuestion(question=question, applicable=False)
                for question in SAFETY_PERMIT_QUESTIONS
            ),
            permit_required=False,
        ),
    )

    workbook = load_workbook(BytesIO(render_work_order_xlsx(content)))

    assert workbook["01_현장확인"].sheet_state == "visible"
    assert workbook["03_펌프점검"].sheet_state == "hidden"
    assert workbook["04_열교환기점검"].sheet_state == "hidden"
